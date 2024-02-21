import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Translate column 1 to a letter
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print('max column for example: %s' % get_column_letter(sheet.max_column))
# Get A's number
print('column index for A: %s' % column_index_from_string('A'))
print('column index for AA: %s' % column_index_from_string('AA'))
