import openpyxl

# Create a blank workbook
wb = openpyxl.Workbook()

print(wb.sheetnames)

sheet = wb.active
print(sheet.title)

# Change title
sheet.title = 'MySheet'
print(wb.sheetnames)

wb.save("blank_workbook.xlsx")