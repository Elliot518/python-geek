import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

# merge all these cells
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'

# merge these two cells
sheet.merge_cells('C5:D5')

sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')
print("Workbook Saved!")
