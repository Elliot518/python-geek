import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

# get a cell value from the sheet
print(sheet['A1'].value)

# get another cell value from the sheet
c = sheet['B1']
print(c.value)

# Get the row, column, and value from the cell
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))

# you can also get a cell(B1) using the sheet’s cell() method and passing integers for its row and column
print(sheet.cell(row=1, column=2).value)
# Go through every other row
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
