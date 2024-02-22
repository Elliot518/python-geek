import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)
# Add a new sheet
wb.create_sheet()
print(wb.sheetnames)
# Create a new sheet at index 0
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)

del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)

# writing Values to Cells
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!'
print(sheet['A1'].value)
wb.save("new_workbook.xlsx")

